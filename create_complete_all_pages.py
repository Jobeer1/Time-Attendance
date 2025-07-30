import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os

def create_complete_promed_report():
    """Create a complete Excel file with all three pages of PROMED reports"""
    try:
        # File paths
        csv_files = [
            r"c:\Users\Admin\Desktop\ELC\Workspaces\Time Attendance\reports1.csv",
            r"c:\Users\Admin\Desktop\ELC\Workspaces\Time Attendance\reports2.csv",
            r"c:\Users\Admin\Desktop\ELC\Workspaces\Time Attendance\reports3.csv"
        ]
        excel_file = r"c:\Users\Admin\Desktop\ELC\Workspaces\Time Attendance\PROMED_Complete_All_Pages.xlsx"
        
        print("Creating comprehensive PROMED report with all three pages...")
        
        def process_csv_file(csv_file):
            """Process a single CSV file and return header info and data"""
            with open(csv_file, 'r', encoding='utf-8') as f:
                lines = f.readlines()
            
            # Extract header information (first 6 lines)
            header_info = []
            for i in range(6):
                parts = lines[i].strip().split(',', 1)
                if len(parts) == 2:
                    header_info.append([parts[0], parts[1]])
                else:
                    header_info.append([parts[0], ''])
            
            # Extract data starting from line 8
            data_lines = lines[7:]
            data_content = ''.join(data_lines)
            
            from io import StringIO
            data_df = pd.read_csv(StringIO(data_content))
            
            return header_info, data_df
        
        # Process all files
        all_data = []
        all_headers = []
        
        for i, csv_file in enumerate(csv_files):
            print(f"Processing {csv_file}...")
            header, data = process_csv_file(csv_file)
            all_headers.append(header)
            all_data.append(data)
            print(f"Page {i+1} records: {len(data)}")
        
        # Combine all data
        combined_data = pd.concat(all_data, ignore_index=True)
        total_records = len(combined_data)
        
        print(f"Total combined records: {total_records}")
        
        # Create Excel workbook
        wb = openpyxl.Workbook()
        
        # Define styles
        title_font = Font(name='Arial', size=16, bold=True, color='000080')
        subtitle_font = Font(name='Arial', size=14, bold=True, color='000080')
        header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        data_font = Font(name='Arial', size=10)
        alt_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        summary_font = Font(name='Arial', size=11, bold=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Create Executive Summary sheet
        summary_sheet = wb.active
        summary_sheet.title = "Executive Summary"
        
        summary_sheet['A1'] = 'PROMED MEDICAL BILLING REPORT'
        summary_sheet['A1'].font = title_font
        summary_sheet.merge_cells('A1:F1')
        summary_sheet['A1'].alignment = Alignment(horizontal='center')
        
        summary_sheet['A2'] = 'COMPLETE BAD DEBT ANALYSIS'
        summary_sheet['A2'].font = subtitle_font
        summary_sheet.merge_cells('A2:F2')
        summary_sheet['A2'].alignment = Alignment(horizontal='center')
        
        # Add report information
        row = 4
        summary_sheet[f'A{row}'] = 'Report Information'
        summary_sheet[f'A{row}'].font = subtitle_font
        row += 1
        
        for key, value in all_headers[0]:
            summary_sheet[f'A{row}'] = key.replace('_', ' ').title()
            summary_sheet[f'B{row}'] = value
            summary_sheet[f'A{row}'].font = summary_font
            row += 1
        
        # Add summary statistics
        row += 1
        summary_sheet[f'A{row}'] = 'COMPREHENSIVE STATISTICS'
        summary_sheet[f'A{row}'].font = subtitle_font
        row += 1
        
        summary_sheet[f'A{row}'] = 'Total Records (All Pages):'
        summary_sheet[f'B{row}'] = total_records
        summary_sheet[f'A{row}'].font = summary_font
        row += 1
        
        for i, data in enumerate(all_data, 1):
            summary_sheet[f'A{row}'] = f'Page {i} Records:'
            summary_sheet[f'B{row}'] = len(data)
            summary_sheet[f'A{row}'].font = summary_font
            row += 1
        
        # Calculate financial totals
        try:
            total_outstanding = pd.to_numeric(combined_data.iloc[:, 10], errors='coerce').sum()
            total_current = pd.to_numeric(combined_data.iloc[:, 4], errors='coerce').sum()
            total_30_days = pd.to_numeric(combined_data.iloc[:, 5], errors='coerce').sum()
            total_60_days = pd.to_numeric(combined_data.iloc[:, 6], errors='coerce').sum()
            total_90_days = pd.to_numeric(combined_data.iloc[:, 7], errors='coerce').sum()
            total_120_days = pd.to_numeric(combined_data.iloc[:, 8], errors='coerce').sum()
            total_150_days = pd.to_numeric(combined_data.iloc[:, 9], errors='coerce').sum()
            
            row += 1
            summary_sheet[f'A{row}'] = 'FINANCIAL ANALYSIS'
            summary_sheet[f'A{row}'].font = subtitle_font
            row += 1
            
            financial_data = [
                ('Total Outstanding Amount:', total_outstanding),
                ('Current Amount:', total_current),
                ('30 Days Outstanding:', total_30_days),
                ('60 Days Outstanding:', total_60_days),
                ('90 Days Outstanding:', total_90_days),
                ('120 Days Outstanding:', total_120_days),
                ('150+ Days Outstanding:', total_150_days)
            ]
            
            for label, amount in financial_data:
                summary_sheet[f'A{row}'] = label
                summary_sheet[f'B{row}'] = amount
                summary_sheet[f'B{row}'].number_format = '#,##0.00'
                summary_sheet[f'A{row}'].font = summary_font
                row += 1
                
        except Exception as e:
            print(f"Warning: Could not calculate financial totals: {e}")
        
        # Adjust column widths
        summary_sheet.column_dimensions['A'].width = 30
        summary_sheet.column_dimensions['B'].width = 20
        
        # Create Combined Data sheet
        data_sheet = wb.create_sheet(title="All Patient Data")
        
        data_sheet['A1'] = 'COMPLETE PATIENT BILLING DATA (ALL PAGES)'
        data_sheet['A1'].font = title_font
        data_sheet.merge_cells('A1:Q1')
        data_sheet['A1'].alignment = Alignment(horizontal='center')
        
        # Add column headers
        start_row = 3
        for col_num, column_title in enumerate(combined_data.columns, 1):
            cell = data_sheet.cell(row=start_row, column=col_num)
            cell.value = column_title.replace('_', ' ').title()
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Add data rows
        for row_num, (index, row) in enumerate(combined_data.iterrows(), start_row + 1):
            for col_num, value in enumerate(row, 1):
                cell = data_sheet.cell(row=row_num, column=col_num)
                
                # Handle numeric values for amount columns
                if col_num in [4, 5, 6, 7, 8, 9, 10]:  # Amount columns
                    try:
                        if pd.notna(value) and str(value).strip() != '':
                            cell.value = float(value)
                            cell.number_format = '#,##0.00'
                        else:
                            cell.value = 0.0
                            cell.number_format = '#,##0.00'
                    except:
                        cell.value = str(value) if pd.notna(value) else ''
                else:
                    cell.value = str(value) if pd.notna(value) else ''
                
                cell.font = data_font
                cell.alignment = Alignment(vertical='center')
                cell.border = border
                
                # Alternate row coloring
                if row_num % 2 == 0:
                    cell.fill = alt_fill
        
        # Auto-adjust column widths
        for col_num in range(1, len(combined_data.columns) + 1):
            max_length = 0
            column_letter = openpyxl.utils.get_column_letter(col_num)
            
            for row_num in range(start_row, data_sheet.max_row + 1):
                cell = data_sheet.cell(row=row_num, column=col_num)
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            data_sheet.column_dimensions[column_letter].width = adjusted_width
        
        # Add auto-filter
        last_col_letter = openpyxl.utils.get_column_letter(len(combined_data.columns))
        filter_range = f"A{start_row}:{last_col_letter}{data_sheet.max_row}"
        data_sheet.auto_filter.ref = filter_range
        
        # Freeze panes
        data_sheet.freeze_panes = f'A{start_row + 1}'
        
        # Create separate sheets for each page
        for page_num, (page_data, page_header) in enumerate(zip(all_data, all_headers), 1):
            sheet = wb.create_sheet(title=f"Page {page_num} Data")
            
            sheet['A1'] = f'PATIENT BILLING DATA - PAGE {page_num}'
            sheet['A1'].font = subtitle_font
            sheet.merge_cells('A1:Q1')
            sheet['A1'].alignment = Alignment(horizontal='center')
            
            # Add headers and data
            start_row = 3
            for col_num, column_title in enumerate(page_data.columns, 1):
                cell = sheet.cell(row=start_row, column=col_num)
                cell.value = column_title.replace('_', ' ').title()
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
            
            for row_num, (index, row) in enumerate(page_data.iterrows(), start_row + 1):
                for col_num, value in enumerate(row, 1):
                    cell = sheet.cell(row=row_num, column=col_num)
                    
                    if col_num in [4, 5, 6, 7, 8, 9, 10]:
                        try:
                            if pd.notna(value) and str(value).strip() != '':
                                cell.value = float(value)
                                cell.number_format = '#,##0.00'
                            else:
                                cell.value = 0.0
                                cell.number_format = '#,##0.00'
                        except:
                            cell.value = str(value) if pd.notna(value) else ''
                    else:
                        cell.value = str(value) if pd.notna(value) else ''
                    
                    cell.font = data_font
                    cell.alignment = Alignment(vertical='center')
                    cell.border = border
                    
                    if row_num % 2 == 0:
                        cell.fill = alt_fill
        
        # Save the workbook
        wb.save(excel_file)
        print(f"\n✅ Complete Excel file created successfully: {excel_file}")
        
        return True
        
    except Exception as e:
        print(f"❌ Error during conversion: {str(e)}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    success = create_complete_promed_report()
    if success:
        print("\n🎉 PROMED Complete Report with All Pages Created!")
        print("📄 File: PROMED_Complete_All_Pages.xlsx")
        print("📋 Contains 5 sheets:")
        print("   1. Executive Summary - Complete overview and financial analysis")
        print("   2. All Patient Data - Combined data from all 3 pages")
        print("   3. Page 1 Data - Individual page 1 data")
        print("   4. Page 2 Data - Individual page 2 data")
        print("   5. Page 3 Data - Individual page 3 data")
        print("\n💼 Professional Features:")
        print("   ✓ Executive summary with financial analysis")
        print("   ✓ Complete aging analysis (Current, 30, 60, 90, 120, 150+ days)")
        print("   ✓ Professional formatting and styling")
        print("   ✓ Auto-filters for advanced data analysis")
        print("   ✓ Frozen header rows for easy navigation")
        print("   ✓ Currency formatting for all amounts")
        print("   ✓ Alternating row colors for readability")
        print("   ✓ Comprehensive summary statistics")
        print("   ✓ Professional borders and presentation")
    else:
        print("❌ Failed to create complete report!")

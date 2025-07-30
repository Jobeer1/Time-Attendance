import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import os

def convert_csv_to_excel():
    # Read the CSV file
    csv_file_path = r"c:\Users\Admin\Desktop\ELC\Workspaces\Time Attendance\reports1.csv"
    excel_file_path = r"c:\Users\Admin\Desktop\ELC\Workspaces\Time Attendance\reports1.xlsx"
    
    # Read CSV file
    df = pd.read_csv(csv_file_path)
    
    # Create a new workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PROMED Medical Report"
    
    # Define styles
    header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    
    data_font = Font(name='Arial', size=10)
    alt_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    
    title_font = Font(name='Arial', size=14, bold=True, color='000080')
    subtitle_font = Font(name='Arial', size=11, bold=True)
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Add report header information
    ws['A1'] = 'PROMED MEDICAL BILLING REPORT'
    ws['A1'].font = title_font
    ws.merge_cells('A1:Q1')
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Extract header information from the first 6 rows
    header_info = df.head(6)
    
    row_num = 3
    for index, row in header_info.iterrows():
        ws[f'A{row_num}'] = row.iloc[0]
        ws[f'B{row_num}'] = row.iloc[1]
        ws[f'A{row_num}'].font = subtitle_font
        row_num += 1
    
    # Skip the empty row and get the actual data
    data_df = df.iloc[7:].reset_index(drop=True)
    
    # Add column headers starting from row 10
    start_row = 10
    for col_num, column_title in enumerate(data_df.columns, 1):
        cell = ws.cell(row=start_row, column=col_num)
        cell.value = column_title.replace('_', ' ').title()
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Add data rows
    for row_num, (index, row) in enumerate(data_df.iterrows(), start_row + 1):
        for col_num, value in enumerate(row, 1):
            cell = ws.cell(row=row_num, column=col_num)
            
            # Handle numeric values
            if col_num in [4, 5, 6, 7, 8, 9, 10]:  # Amount columns
                try:
                    cell.value = float(value) if pd.notna(value) and value != '' else 0.0
                    cell.number_format = '#,##0.00'
                except:
                    cell.value = value
            else:
                cell.value = value if pd.notna(value) else ''
            
            cell.font = data_font
            cell.border = border
            cell.alignment = Alignment(vertical='center')
            
            # Alternate row coloring
            if row_num % 2 == 0:
                cell.fill = alt_fill
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Add summary section
    last_row = ws.max_row + 2
    
    # Calculate totals for numeric columns
    ws[f'A{last_row}'] = 'SUMMARY'
    ws[f'A{last_row}'].font = title_font
    
    summary_row = last_row + 1
    ws[f'A{summary_row}'] = 'Total Records:'
    ws[f'B{summary_row}'] = len(data_df)
    ws[f'A{summary_row}'].font = subtitle_font
    
    # Calculate total outstanding amount
    try:
        total_outstanding = data_df.iloc[:, 10].astype(float).sum()  # Outstanding column
        ws[f'A{summary_row + 1}'] = 'Total Outstanding:'
        ws[f'B{summary_row + 1}'] = total_outstanding
        ws[f'B{summary_row + 1}'].number_format = '#,##0.00'
        ws[f'A{summary_row + 1}'].font = subtitle_font
    except:
        pass
    
    # Add filters to the data
    ws.auto_filter.ref = f"A{start_row}:{ws.max_column}{ws.max_row - 3}"
    
    # Freeze panes at the header row
    ws.freeze_panes = f'A{start_row + 1}'
    
    # Save the workbook
    wb.save(excel_file_path)
    print(f"Excel file created successfully: {excel_file_path}")
    
    return excel_file_path

if __name__ == "__main__":
    convert_csv_to_excel()

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os

def convert_reports_to_excel():
    try:
        # File paths
        csv_file = r"c:\Users\Admin\Desktop\ELC\Workspaces\Time Attendance\reports1.csv"
        excel_file = r"c:\Users\Admin\Desktop\ELC\Workspaces\Time Attendance\PROMED_Medical_Report.xlsx"
        
        print(f"Processing CSV file: {csv_file}")
        
        # Read the file line by line to handle the mixed structure
        with open(csv_file, 'r', encoding='utf-8') as f:
            lines = f.readlines()
        
        # Extract header information (first 6 lines)
        header_info = []
        for i in range(6):
            parts = lines[i].strip().split(',', 1)
            if len(parts) == 2:
                header_info.append([parts[0], parts[1]])
            else:
                header_info.append([parts[0], ''])
        
        # Extract data starting from line 8 (skip the blank line at 7)
        data_lines = lines[7:]  # This includes the column headers
        
        # Create a temporary CSV content for the data part
        data_content = ''.join(data_lines)
        
        # Read the data part into a DataFrame
        from io import StringIO
        data_df = pd.read_csv(StringIO(data_content))
        
        print(f"Header info extracted: {len(header_info)} rows")
        print(f"Data extracted: {len(data_df)} rows")
        
        # Create Excel workbook
        wb = openpyxl.Workbook()
        
        # Create Report Info sheet
        info_sheet = wb.active
        info_sheet.title = "Report Information"
        
        # Define styles
        title_font = Font(name='Arial', size=14, bold=True, color='000080')
        header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        data_font = Font(name='Arial', size=10)
        alt_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        
        # Add title
        info_sheet['A1'] = 'PROMED MEDICAL BILLING REPORT'
        info_sheet['A1'].font = title_font
        info_sheet.merge_cells('A1:B1')
        info_sheet['A1'].alignment = Alignment(horizontal='center')
        
        # Add header information
        for i, (key, value) in enumerate(header_info, start=3):
            info_sheet[f'A{i}'] = key.replace('_', ' ').title()
            info_sheet[f'B{i}'] = value
            info_sheet[f'A{i}'].font = Font(name='Arial', size=11, bold=True)
        
        # Adjust column widths for info sheet
        info_sheet.column_dimensions['A'].width = 25
        info_sheet.column_dimensions['B'].width = 40
        
        # Create Patient Data sheet
        data_sheet = wb.create_sheet(title="Patient Data")
        
        # Add title to data sheet
        data_sheet['A1'] = 'PATIENT BILLING DATA'
        data_sheet['A1'].font = title_font
        data_sheet.merge_cells('A1:Q1')
        data_sheet['A1'].alignment = Alignment(horizontal='center')
        
        # Add column headers starting from row 3
        start_row = 3
        for col_num, column_title in enumerate(data_df.columns, 1):
            cell = data_sheet.cell(row=start_row, column=col_num)
            cell.value = column_title.replace('_', ' ').title()
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Add data rows
        for row_num, (index, row) in enumerate(data_df.iterrows(), start_row + 1):
            for col_num, value in enumerate(row, 1):
                cell = data_sheet.cell(row=row_num, column=col_num)
                
                # Handle numeric values for amount columns
                if col_num in [4, 5, 6, 7, 8, 9, 10]:  # Amount columns
                    try:
                        if pd.notna(value) and str(value).strip() != '':
                            cell.value = float(value)
                            cell.number_format = '#,##0.00'
                        else:
                            cell.value = 0.0
                            cell.number_format = '#,##0.00'
                    except:
                        cell.value = str(value) if pd.notna(value) else ''
                else:
                    cell.value = str(value) if pd.notna(value) else ''
                
                cell.font = data_font
                cell.alignment = Alignment(vertical='center')
                
                # Alternate row coloring
                if row_num % 2 == 0:
                    cell.fill = alt_fill
        
        # Auto-adjust column widths for data sheet
        for col_num in range(1, len(data_df.columns) + 1):
            max_length = 0
            column_letter = openpyxl.utils.get_column_letter(col_num)
            
            # Check all cells in this column
            for row_num in range(start_row, data_sheet.max_row + 1):
                cell = data_sheet.cell(row=row_num, column=col_num)
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            data_sheet.column_dimensions[column_letter].width = adjusted_width
        
        # Add summary information
        last_row = data_sheet.max_row + 2
        data_sheet[f'A{last_row}'] = 'SUMMARY STATISTICS'
        data_sheet[f'A{last_row}'].font = title_font
        
        # Calculate totals
        total_records = len(data_df)
        data_sheet[f'A{last_row + 1}'] = 'Total Records:'
        data_sheet[f'B{last_row + 1}'] = total_records
        
        try:
            # Calculate total outstanding (column K, index 10)
            outstanding_col = data_df.iloc[:, 10]
            total_outstanding = pd.to_numeric(outstanding_col, errors='coerce').sum()
            data_sheet[f'A{last_row + 2}'] = 'Total Outstanding Amount:'
            data_sheet[f'B{last_row + 2}'] = total_outstanding
            data_sheet[f'B{last_row + 2}'].number_format = '#,##0.00'
        except:
            pass
        
        # Add auto-filter to data
        last_col_letter = openpyxl.utils.get_column_letter(len(data_df.columns))
        filter_range = f"A{start_row}:{last_col_letter}{data_sheet.max_row - 3}"
        data_sheet.auto_filter.ref = filter_range
        
        # Freeze panes
        data_sheet.freeze_panes = f'A{start_row + 1}'
        
        # Save the workbook
        wb.save(excel_file)
        print(f"Excel file created successfully: {excel_file}")
        
        # Print summary
        print(f"\nSummary:")
        print(f"- Total patient records: {total_records}")
        print(f"- Report date: 25.06.25")
        print(f"- Doctor: DR. C.I. STOYANOV")
        print(f"- Medical Aid: BAD - BAD DEBT")
        
        return True
        
    except Exception as e:
        print(f"Error during conversion: {str(e)}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    success = convert_reports_to_excel()
    if success:
        print("\n✅ Excel conversion completed successfully!")
        print("📄 File saved as: PROMED_Medical_Report.xlsx")
        print("📊 The file contains two sheets:")
        print("   - Report Information: Header details")
        print("   - Patient Data: Complete patient billing data with formatting")
    else:
        print("❌ Conversion failed! Check the error messages above.")

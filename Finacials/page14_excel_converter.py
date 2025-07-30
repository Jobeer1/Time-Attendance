import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, NamedStyle, Border, Side, numbers

# Read CSV
input_file = 'page14.csv'
df = pd.read_csv(input_file)

# Date columns to format
if 'Last_Visit' in df.columns:
    df['Last_Visit'] = pd.to_datetime(df['Last_Visit'], errors='coerce').dt.strftime('%d/%m/%Y')
if 'Last_Receipt_Payment' in df.columns:
    df['Last_Receipt_Payment'] = pd.to_datetime(df['Last_Receipt_Payment'], errors='coerce').dt.strftime('%d/%m/%Y')

# Currency columns to format
currency_cols = ['Current', '30_Days', '60_Days', '90_Days', '120_Days', '150_Days', 'Outstanding']
for col in currency_cols:
    if col in df.columns:
        df[col] = pd.to_numeric(df[col], errors='coerce')

# Reorder columns for user-friendly view
order = ['Account_ID','Account_Name','Account_Type','Last_Visit','Last_Receipt_Payment','Current','30_Days','60_Days','90_Days','120_Days','150_Days','Outstanding','Status','ID_Number','Claim']
df = df.reindex(columns=order)

# Write to Excel
excel_file = 'page14_corrected.xlsx'
df.to_excel(excel_file, index=False)

# Post-process Excel for formatting
wb = load_workbook(excel_file)
ws = wb.active

# Bold headers
for cell in ws[1]:
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    cell.alignment = Alignment(horizontal='center')

# Rand currency formatting
rand_fmt = 'R #,##0.00'
for col in currency_cols:
    idx = order.index(col) + 1
    for row in ws.iter_rows(min_row=2, min_col=idx, max_col=idx, max_row=ws.max_row):
        for cell in row:
            cell.number_format = rand_fmt
            cell.alignment = Alignment(horizontal='right')

# Date formatting
for col in ['Last_Visit', 'Last_Receipt_Payment']:
    idx = order.index(col) + 1
    for row in ws.iter_rows(min_row=2, min_col=idx, max_col=idx, max_row=ws.max_row):
        for cell in row:
            cell.number_format = 'DD/MM/YYYY'
            cell.alignment = Alignment(horizontal='center')

# Save
wb.save(excel_file)
print(f"Excel file '{excel_file}' created with formatting.")

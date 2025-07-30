# --- HUMAN-FRIENDLY EXTRACTION LOGIC ---
# This script is designed to extract financial account data from messy OCR/CSV files and output a professional, human-friendly Excel file.
#
# - A South African ID number is exactly 13 digits (e.g., 8001015009087).
# - An account number is typically 8-10 digits, but not a 13-digit ID.
# - The script uses regex to find and distinguish these numbers.
# - The output Excel will always have these columns, in this order:
#   Account Name | Account Number | ID Number | Date | Amount (ZAR)
# - Each account is a single row, with all data neatly aligned in columns for easy human reading and financial work.

import pandas as pd
import re
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime

def is_id(val):
    # South African ID: 13 digits
    return bool(re.match(r'\b\d{13}\b', val)) or val.strip().startswith('ID:')

def is_account(val):
    # Account number: 8-10 digits, but not a 13-digit ID
    return bool(re.match(r'\b\d{8,10}\b', val)) and not is_id(val)

def is_amount(val):
    # Accepts numbers with comma or dot as decimal, optional minus, optional R
    return bool(re.match(r'^(R ?)?[\d.,\-]+$', val)) and (',' in val or '.' in val)

def is_date(val):
    # Accepts DD/MM/YYYY, DD-MM-YYYY, YYYY-MM-DD, etc.
    return bool(re.match(r'(\d{1,2}[./-]\d{1,2}[./-]\d{2,4})|(\d{4}-\d{2}-\d{2})', val))

def extract_accounts(csv_path):
    with open(csv_path, encoding='utf-8') as f:
        lines = [line.strip() for line in f if line.strip()]
    accounts = []
    prev_name = ''
    prev_date = ''
    for i, line in enumerate(lines):
        row = [c.replace('"','').strip() for c in line.split(',')]
        # Skip header and address lines
        if any(h in row[0].upper() for h in ["PROMED", "REPORT", "ACCOUNT", "MEDICAL", "LAST RECEIPT PAYMENT", "PO BOX", "STREET", "RESERVE", "MEWS", "BLOCK", "P/BAG", "UNIT", "UNSELECTED"]):
            continue
        id_number = next((c.replace('ID:', '').strip() for c in row if is_id(c)), '')
        account_number = next((c for c in row if is_account(c)), '')
        amount = next((c for c in row if is_amount(c)), '')
        date = next((c for c in row if is_date(c)), '')
        # Name: first non-empty, non-ID, non-amount, non-date, non-account cell
        name = next((c for c in row if c and not is_id(c) and not is_account(c) and not is_amount(c) and not is_date(c)), '')
        # Carry forward previous name/date if missing
        if not name and prev_name:
            name = prev_name
        if not date and prev_date:
            date = prev_date
        # Clean amount to ZAR format
        if amount:
            amount_clean = re.sub(r'[^\d.,-]', '', amount.replace(' ', ''))
            amount_clean = amount_clean.replace(',', '.')
            try:
                amount_float = float(re.sub(r'[^\d.-]', '', amount_clean))
                amount_zar = f"R {amount_float:,.2f}"
            except Exception:
                amount_zar = amount
        else:
            amount_zar = ''
        # Only output rows with at least account number or ID and amount
        if (account_number or id_number) and amount_zar:
            accounts.append({
                'Account Name': name,
                'Account Number': account_number,
                'ID Number': id_number,
                'Date': date,
                'Amount (ZAR)': amount_zar
            })
        # Save name/date for next row if found
        if name:
            prev_name = name
        if date:
            prev_date = date
    df = pd.DataFrame(accounts, columns=['Account Name', 'Account Number', 'ID Number', 'Date', 'Amount (ZAR)'])
    return df

def save_to_xlsx(df, out_path):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Accounts'
    # Title
    ws.merge_cells('A1:E1')
    ws['A1'] = 'PROMED FINANCIAL REPORT - CLEANED & FORMATTED'
    ws['A1'].font = Font(size=14, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')
    # Subtitle
    ws.merge_cells('A2:E2')
    ws['A2'] = f'Report generated: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}'
    ws['A2'].alignment = Alignment(horizontal='center')
    # Header
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    for c_idx, col in enumerate(df.columns, 1):
        cell = ws.cell(row=3, column=c_idx, value=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    # Data
    for r_idx, row in enumerate(df.itertuples(index=False), 4):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)
    # Auto-width
    from openpyxl.cell.cell import MergedCell
    for col in ws.columns:
        max_length = 0
        col_letter = None
        for cell in col:
            if isinstance(cell, MergedCell):
                continue
            if col_letter is None:
                col_letter = cell.column_letter
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        if col_letter:
            ws.column_dimensions[col_letter].width = max_length + 2
    wb.save(out_path)
    print(f'Saved cleaned XLSX: {out_path}')

if __name__ == '__main__':
    import sys
    if len(sys.argv) < 2:
        print('Usage: python extract_accounts_to_xlsx.py <input.csv> [output.xlsx]')
        sys.exit(1)
    csv_path = sys.argv[1]
    out_path = sys.argv[2] if len(sys.argv) > 2 else 'accounts_cleaned.xlsx'
    df = extract_accounts(csv_path)
    if df.empty:
        print('No account data found.')
    else:
        save_to_xlsx(df, out_path)
        print('Done. Please check the output XLSX for a professional, human-friendly layout.')

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime

# Load data
df = pd.read_csv('page11.csv')

# Reorder columns for best readability
ordered_cols = [
    'Account_ID', 'Account_Name', 'Account_Type', 'ID_Number', 'Claim',
    'Last_Visit', 'Last_Receipt_Payment',
    'Current', '30_Days', '60_Days', '90_Days', '120_Days', '150_Days',
    'Outstanding', 'Status'
]
df = df[ordered_cols]

# Format date columns as DD/MM/YYYY
for col in ['Last_Visit', 'Last_Receipt_Payment']:
    df[col] = pd.to_datetime(df[col], errors='coerce').dt.strftime('%d/%m/%Y')

# Create workbook and worksheet
wb = Workbook()
ws = wb.active
ws.title = 'Page 11 Write Offs'

# Title
ws.merge_cells('A1:O1')
ws['A1'] = 'PROMED JUN QUARTERLY WRITE OFFS - PAGE 11'
ws['A1'].font = Font(size=14, bold=True)
ws['A1'].alignment = Alignment(horizontal='center')

# Subtitle
ws.merge_cells('A2:O2')
ws['A2'] = f'Report generated: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}'
ws['A2'].alignment = Alignment(horizontal='center')

# Write DataFrame to worksheet
header_font = Font(bold=True)
header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 4):
    for c_idx, value in enumerate(row, 1):
        cell = ws.cell(row=r_idx, column=c_idx, value=value)
        if r_idx == 4:
            cell.font = header_font
            cell.fill = header_fill

# Format currency columns
currency_cols = ['Current', '30_Days', '60_Days', '90_Days', '120_Days', '150_Days', 'Outstanding']
col_letter_map = {col: idx+1 for idx, col in enumerate(df.columns)}
zar_format = 'R #,##0.00'
for col in currency_cols:
    col_idx = col_letter_map[col]
    for row in range(5, 5 + len(df)):
        ws.cell(row=row, column=col_idx).number_format = zar_format

# Auto-width columns (skip merged cells)
from openpyxl.cell.cell import MergedCell
for col in ws.columns:
    max_length = 0
    col_letter = None
    for cell in col:
        if not isinstance(cell, MergedCell):
            if col_letter is None:
                col_letter = cell.column_letter
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
    if col_letter:
        ws.column_dimensions[col_letter].width = max_length + 2

# Summary section
summary_row = len(df) + 6
ws[f'A{summary_row}'] = 'Summary'
ws[f'A{summary_row}'].font = Font(bold=True)
ws[f'A{summary_row+1}'] = 'Total Accounts:'
ws[f'B{summary_row+1}'] = len(df)
ws[f'A{summary_row+2}'] = 'Total Outstanding:'
ws[f'B{summary_row+2}'] = f"R {df['Outstanding'].sum():,.2f}"
ws[f'A{summary_row+3}'] = 'Average Outstanding:'
ws[f'B{summary_row+3}'] = f"R {df['Outstanding'].mean():,.2f}"
ws[f'A{summary_row+4}'] = 'Highest Outstanding:'
ws[f'B{summary_row+4}'] = f"R {df['Outstanding'].max():,.2f}"
ws[f'A{summary_row+5}'] = 'Lowest Outstanding:'
ws[f'B{summary_row+5}'] = f"R {df['Outstanding'].min():,.2f}"

# Save workbook
wb.save('page11_corrected.xlsx')
print('Excel report generated: page11_corrected.xlsx')

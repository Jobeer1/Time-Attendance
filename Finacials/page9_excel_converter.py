import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime

# Load data
df = pd.read_csv('page9.csv')

# Create workbook and worksheet
wb = Workbook()
ws = wb.active
ws.title = 'Page 9 Write Offs'

# Header
header = list(df.columns)
ws.append(header)
for col in range(1, len(header)+1):
    ws.cell(row=1, column=col).font = Font(bold=True)
    ws.cell(row=1, column=col).fill = PatternFill(start_color='B7DEE8', end_color='B7DEE8', fill_type='solid')
    ws.cell(row=1, column=col).alignment = Alignment(horizontal='center')

# Data rows
for row in df.itertuples(index=False):
    ws.append(row)

# Auto width
for col in ws.columns:
    max_length = 0
    col_letter = col[0].column_letter
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    ws.column_dimensions[col_letter].width = max_length + 2

# Summary section
summary_row = ws.max_row + 2
ws.cell(row=summary_row, column=1, value='Summary').font = Font(bold=True, size=12)
ws.cell(row=summary_row+1, column=1, value='Total Outstanding:')
ws.cell(row=summary_row+1, column=2, value=f"R {df['Outstanding'].sum():,.2f}")
ws.cell(row=summary_row+2, column=1, value='Total Accounts:')
ws.cell(row=summary_row+2, column=2, value=len(df))
ws.cell(row=summary_row+3, column=1, value='Average Outstanding:')
ws.cell(row=summary_row+3, column=2, value=f"R {df['Outstanding'].mean():,.2f}")
ws.cell(row=summary_row+4, column=1, value='Highest Outstanding:')
ws.cell(row=summary_row+4, column=2, value=f"R {df['Outstanding'].max():,.2f}")
ws.cell(row=summary_row+5, column=1, value='Lowest Outstanding:')
ws.cell(row=summary_row+5, column=2, value=f"R {df['Outstanding'].min():,.2f}")

# Save Excel
wb.save('page9_corrected.xlsx')
print('Excel report generated: page9_corrected.xlsx')

#!/usr/bin/env python3
"""
Page 8 Excel Converter - Medical Aid Write-off Data
Converts page8.csv to a professional Excel file with multiple sheets
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
import os
from datetime import datetime

def load_and_process_data(csv_file):
    """Load and process the CSV data"""
    try:
        # Read the CSV file, skipping the first 7 rows (6 metadata + 1 empty row)
        # The column headers are on row 8 (index 7)
        df = pd.read_csv(csv_file, skiprows=7)
        
        # Data is already properly loaded with headers
        data_df = df.copy()
        
        # Reset index
        data_df = data_df.reset_index(drop=True)
        
        # Convert numeric columns
        numeric_columns = ['Current', '30_Days', '60_Days', '90_Days', '120_Days', '150_Days', 'Outstanding']
        for col in numeric_columns:
            if col in data_df.columns:
                data_df[col] = pd.to_numeric(data_df[col], errors='coerce').fillna(0)
        
        # Convert Last_Receipt_Payment to numeric
        if 'Last_Receipt_Payment' in data_df.columns:
            data_df['Last_Receipt_Payment'] = pd.to_numeric(data_df['Last_Receipt_Payment'], errors='coerce').fillna(0)
        
        # Parse dates
        if 'Last_Visit' in data_df.columns:
            data_df['Last_Visit'] = pd.to_datetime(data_df['Last_Visit'], format='%d.%m.%y', errors='coerce')
        
        print(f"Loaded {len(data_df)} records from {csv_file}")
        return data_df
        
    except Exception as e:
        print(f"Error loading data: {e}")
        return None

def get_report_metadata(csv_file):
    """Extract report metadata from CSV"""
    try:
        with open(csv_file, 'r') as f:
            lines = f.readlines()
        
        metadata = {}
        for i, line in enumerate(lines[:6]):  # First 6 lines contain metadata
            if ',' in line:
                key, value = line.strip().split(',', 1)
                metadata[key] = value
        
        return metadata
    except Exception as e:
        print(f"Error reading metadata: {e}")
        return {}

def create_summary_stats(df):
    """Create summary statistics"""
    if df is None or df.empty:
        return {}
    
    numeric_columns = ['Current', '30_Days', '60_Days', '90_Days', '120_Days', '150_Days', 'Outstanding']
    
    summary = {
        'total_accounts': len(df),
        'total_outstanding': df['Outstanding'].sum() if 'Outstanding' in df.columns else 0,
        'total_current': df['Current'].sum() if 'Current' in df.columns else 0,
        'avg_outstanding': df['Outstanding'].mean() if 'Outstanding' in df.columns else 0,
        'max_outstanding': df['Outstanding'].max() if 'Outstanding' in df.columns else 0,
        'min_outstanding': df['Outstanding'].min() if 'Outstanding' in df.columns else 0,
    }
    
    # Aging analysis
    aging_totals = {}
    for col in numeric_columns:
        if col in df.columns:
            aging_totals[col] = df[col].sum()
    
    summary['aging_totals'] = aging_totals
    
    return summary

def create_aging_analysis(df):
    """Create aging analysis breakdown"""
    if df is None or df.empty:
        return pd.DataFrame()
    
    aging_columns = ['Current', '30_Days', '60_Days', '90_Days', '120_Days', '150_Days']
    aging_data = []
    
    for col in aging_columns:
        if col in df.columns:
            total = df[col].sum()
            count = (df[col] > 0).sum()
            avg = df[col].mean() if count > 0 else 0
            aging_data.append({
                'Age_Bracket': col.replace('_', ' '),
                'Total_Amount': total,
                'Account_Count': count,
                'Average_Amount': avg,
                'Percentage_of_Total': (total / df['Outstanding'].sum() * 100) if df['Outstanding'].sum() > 0 else 0
            })
    
    return pd.DataFrame(aging_data)

def get_top_accounts(df, n=10):
    """Get top N accounts by outstanding amount"""
    if df is None or df.empty:
        return pd.DataFrame()
    
    top_accounts = df.nlargest(n, 'Outstanding')[
        ['Account_ID', 'Account_Name', 'Outstanding', 'Current', 'Last_Visit', 'Status']
    ].copy()
    
    return top_accounts

def style_header(ws, start_row, end_row, start_col, end_col):
    """Apply header styling"""
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

def style_data_sheet(ws, df):
    """Apply styling to data sheet"""
    # Header styling
    style_header(ws, 1, 1, 1, len(df.columns))
    
    # Data styling
    for row in range(2, len(df) + 2):
        for col in range(1, len(df.columns) + 1):
            cell = ws.cell(row=row, column=col)
            
            # Alternate row colors
            if row % 2 == 0:
                cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            
            # Number formatting for currency columns
            if col in [5, 6, 7, 8, 9, 10, 11]:  # Numeric columns
                cell.number_format = '#,##0.00'
            
            # Center alignment for certain columns
            if col in [1, 12, 14]:  # Account_ID, Status, Status_Code
                cell.alignment = Alignment(horizontal="center")

def auto_adjust_columns(ws):
    """Auto-adjust column widths"""
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
        ws.column_dimensions[column_letter].width = adjusted_width

def create_excel_file(df, metadata, output_file):
    """Create the main Excel file with multiple sheets"""
    try:
        # Create workbook
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Sheet 1: Summary
        ws_summary = wb.create_sheet("Summary")
        summary_stats = create_summary_stats(df)
        
        # Add metadata and summary
        current_row = 1
        
        # Report metadata
        ws_summary.cell(row=current_row, column=1, value="MEDICAL AID WRITE-OFF REPORT - PAGE 8")
        ws_summary.cell(row=current_row, column=1).font = Font(size=16, bold=True)
        current_row += 2
        
        for key, value in metadata.items():
            ws_summary.cell(row=current_row, column=1, value=key.replace('_', ' ').title() + ":")
            ws_summary.cell(row=current_row, column=2, value=value)
            ws_summary.cell(row=current_row, column=1).font = Font(bold=True)
            current_row += 1
        
        current_row += 1
        
        # Summary statistics
        ws_summary.cell(row=current_row, column=1, value="SUMMARY STATISTICS")
        ws_summary.cell(row=current_row, column=1).font = Font(size=14, bold=True)
        current_row += 1
        
        summary_items = [
            ("Total Accounts", summary_stats.get('total_accounts', 0)),
            ("Total Outstanding Amount", f"R {summary_stats.get('total_outstanding', 0):,.2f}"),
            ("Total Current Amount", f"R {summary_stats.get('total_current', 0):,.2f}"),
            ("Average Outstanding", f"R {summary_stats.get('avg_outstanding', 0):,.2f}"),
            ("Maximum Outstanding", f"R {summary_stats.get('max_outstanding', 0):,.2f}"),
            ("Minimum Outstanding", f"R {summary_stats.get('min_outstanding', 0):,.2f}"),
        ]
        
        for item, value in summary_items:
            ws_summary.cell(row=current_row, column=1, value=item + ":")
            ws_summary.cell(row=current_row, column=2, value=value)
            ws_summary.cell(row=current_row, column=1).font = Font(bold=True)
            current_row += 1
        
        # Sheet 2: Raw Data
        ws_data = wb.create_sheet("Raw Data")
        
        # Add data to sheet
        for r in dataframe_to_rows(df, index=False, header=True):
            ws_data.append(r)
        
        # Style the data sheet
        style_data_sheet(ws_data, df)
        auto_adjust_columns(ws_data)
        
        # Sheet 3: Aging Analysis
        ws_aging = wb.create_sheet("Aging Analysis")
        aging_df = create_aging_analysis(df)
        
        if not aging_df.empty:
            for r in dataframe_to_rows(aging_df, index=False, header=True):
                ws_aging.append(r)
            
            style_header(ws_aging, 1, 1, 1, len(aging_df.columns))
            auto_adjust_columns(ws_aging)
            
            # Format currency columns
            for row in range(2, len(aging_df) + 2):
                ws_aging.cell(row=row, column=2).number_format = '#,##0.00'  # Total_Amount
                ws_aging.cell(row=row, column=4).number_format = '#,##0.00'  # Average_Amount
                ws_aging.cell(row=row, column=5).number_format = '0.00%'      # Percentage
        
        # Sheet 4: Top Accounts
        ws_top = wb.create_sheet("Top Accounts")
        top_accounts_df = get_top_accounts(df, 10)
        
        if not top_accounts_df.empty:
            for r in dataframe_to_rows(top_accounts_df, index=False, header=True):
                ws_top.append(r)
            
            style_header(ws_top, 1, 1, 1, len(top_accounts_df.columns))
            auto_adjust_columns(ws_top)
            
            # Format currency columns
            for row in range(2, len(top_accounts_df) + 2):
                ws_top.cell(row=row, column=3).number_format = '#,##0.00'  # Outstanding
                ws_top.cell(row=row, column=4).number_format = '#,##0.00'  # Current
        
        # Adjust summary sheet columns
        auto_adjust_columns(ws_summary)
        
        # Save the workbook
        wb.save(output_file)
        print(f"Excel file created successfully: {output_file}")
        
        return True
        
    except Exception as e:
        print(f"Error creating Excel file: {e}")
        return False

def main():
    """Main execution function"""
    # File paths
    csv_file = "page8.csv"
    output_file = "page8_corrected.xlsx"
    
    # Check if input file exists
    if not os.path.exists(csv_file):
        print(f"Error: Input file {csv_file} not found!")
        return
    
    print("Starting Page 8 Excel conversion...")
    print("=" * 50)
    
    # Load data
    df = load_and_process_data(csv_file)
    if df is None:
        print("Failed to load data. Exiting.")
        return
    
    # Get metadata
    metadata = get_report_metadata(csv_file)
    
    # Create Excel file
    success = create_excel_file(df, metadata, output_file)
    
    if success:
        print("=" * 50)
        print("CONVERSION COMPLETED SUCCESSFULLY!")
        print(f"Input file: {csv_file}")
        print(f"Output file: {output_file}")
        print(f"Records processed: {len(df)}")
        print("=" * 50)
        
        # Display summary
        summary = create_summary_stats(df)
        print(f"Total Outstanding Amount: R {summary.get('total_outstanding', 0):,.2f}")
        print(f"Average Outstanding: R {summary.get('avg_outstanding', 0):,.2f}")
        print(f"Number of Accounts: {summary.get('total_accounts', 0)}")
    else:
        print("Conversion failed!")

if __name__ == "__main__":
    main()

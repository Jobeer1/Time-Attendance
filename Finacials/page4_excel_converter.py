#!/usr/bin/env python3
"""
Convert page4.csv (JUN - QUARTERLY WRITE OFF - PAGE 4) to professionally formatted Excel file
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.formatting.rule import ColorScaleRule
import os
from datetime import datetime

def create_page4_excel():
    # Read the CSV file
    csv_path = 'page4.csv'
    if not os.path.exists(csv_path):
        print(f"Error: {csv_path} not found!")
        return
    
    # Load the data with proper handling of mixed data types
    # Read header information separately
    header_data = []
    with open(csv_path, 'r') as f:
        lines = f.readlines()
        # First 6 lines are header info
        for i in range(6):
            if i < len(lines):
                parts = lines[i].strip().split(',', 1)  # Split only on first comma
                if len(parts) >= 2:
                    header_data.append(parts)
                else:
                    header_data.append([parts[0], ''])
    
    # Read the actual data starting from line 8 (skip empty line 7)
    data_df = pd.read_csv(csv_path, skiprows=7, dtype=str, na_filter=False)
    
    # Convert numerical columns
    numerical_cols = ['Last_Receipt_Payment', 'Current', '30_Days', '60_Days', '90_Days', '120_Days', '150_Days', 'Outstanding']
    for col in numerical_cols:
        if col in data_df.columns:
            # Replace empty strings with 0 before converting
            data_df[col] = data_df[col].replace('', '0')
            data_df[col] = pd.to_numeric(data_df[col], errors='coerce').fillna(0)
    
    # Create workbook
    wb = Workbook()
    
    # Create Executive Summary sheet
    exec_ws = wb.active
    exec_ws.title = "Executive Summary"
    
    # Style definitions
    header_font = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    subheader_font = Font(name='Calibri', size=12, bold=True, color='1F4E79')
    normal_font = Font(name='Calibri', size=11)
    currency_font = Font(name='Calibri', size=11, bold=True)
    
    # Executive Summary Header
    exec_ws['A1'] = 'PROMED MEDICAL - QUARTERLY WRITE OFF REPORT - PAGE 4'
    exec_ws['A1'].font = Font(name='Calibri', size=16, bold=True, color='1F4E79')
    exec_ws.merge_cells('A1:F1')
    
    exec_ws['A2'] = f'Generated: {datetime.now().strftime("%B %d, %Y")}'
    exec_ws['A2'].font = Font(name='Calibri', size=10, italic=True)
    exec_ws.merge_cells('A2:F2')
    
    # Report Information
    row = 4
    exec_ws[f'A{row}'] = 'REPORT INFORMATION'
    exec_ws[f'A{row}'].font = subheader_font
    row += 1
    
    for header_row in header_data:
        if len(header_row) >= 2 and header_row[1].strip():
            exec_ws[f'A{row}'] = f'{header_row[0]}:'
            exec_ws[f'B{row}'] = str(header_row[1])
            exec_ws[f'A{row}'].font = normal_font
            exec_ws[f'B{row}'].font = currency_font
            row += 1
    
    row += 1
    
    # Financial Summary
    exec_ws[f'A{row}'] = 'FINANCIAL SUMMARY'
    exec_ws[f'A{row}'].font = subheader_font
    row += 1
    
    # Calculate totals
    total_outstanding = data_df['Outstanding'].sum()
    total_current = data_df['Current'].sum()
    total_30_days = data_df['30_Days'].sum()
    total_60_days = data_df['60_Days'].sum()
    total_90_days = data_df['90_Days'].sum()
    total_120_days = data_df['120_Days'].sum()
    total_150_days = data_df['150_Days'].sum()
    total_accounts = len(data_df)
    avg_outstanding = data_df['Outstanding'].mean()
    total_payments = data_df['Last_Receipt_Payment'].sum()
    
    # Financial metrics
    metrics = [
        ('Total Outstanding Amount', f'R {total_outstanding:,.2f}'),
        ('Total Accounts', f'{total_accounts:,}'),
        ('Average Outstanding per Account', f'R {avg_outstanding:,.2f}'),
        ('Total Receipt Payments', f'R {total_payments:,.2f}'),
        ('Current Amount', f'R {total_current:,.2f}'),
        ('30 Days Overdue', f'R {total_30_days:,.2f}'),
        ('60 Days Overdue', f'R {total_60_days:,.2f}'),
        ('90 Days Overdue', f'R {total_90_days:,.2f}'),
        ('120 Days Overdue', f'R {total_120_days:,.2f}'),
        ('150+ Days Overdue', f'R {total_150_days:,.2f}')
    ]
    
    for metric, value in metrics:
        exec_ws[f'A{row}'] = metric
        exec_ws[f'B{row}'] = value
        exec_ws[f'A{row}'].font = normal_font
        exec_ws[f'B{row}'].font = currency_font
        row += 1
    
    row += 1
    
    # Age Analysis
    exec_ws[f'A{row}'] = 'AGE ANALYSIS'
    exec_ws[f'A{row}'].font = subheader_font
    row += 1
    
    age_analysis = [
        ('Current', total_current, (total_current/total_outstanding*100) if total_outstanding > 0 else 0),
        ('30 Days', total_30_days, (total_30_days/total_outstanding*100) if total_outstanding > 0 else 0),
        ('60 Days', total_60_days, (total_60_days/total_outstanding*100) if total_outstanding > 0 else 0),
        ('90 Days', total_90_days, (total_90_days/total_outstanding*100) if total_outstanding > 0 else 0),
        ('120 Days', total_120_days, (total_120_days/total_outstanding*100) if total_outstanding > 0 else 0),
        ('150+ Days', total_150_days, (total_150_days/total_outstanding*100) if total_outstanding > 0 else 0)
    ]
    
    exec_ws[f'A{row}'] = 'Age Category'
    exec_ws[f'B{row}'] = 'Amount'
    exec_ws[f'C{row}'] = 'Percentage'
    exec_ws[f'A{row}'].font = header_font
    exec_ws[f'B{row}'].font = header_font
    exec_ws[f'C{row}'].font = header_font
    exec_ws[f'A{row}'].fill = header_fill
    exec_ws[f'B{row}'].fill = header_fill
    exec_ws[f'C{row}'].fill = header_fill
    row += 1
    
    for category, amount, percentage in age_analysis:
        exec_ws[f'A{row}'] = category
        exec_ws[f'B{row}'] = f'R {amount:,.2f}'
        exec_ws[f'C{row}'] = f'{percentage:.1f}%'
        exec_ws[f'A{row}'].font = normal_font
        exec_ws[f'B{row}'].font = normal_font
        exec_ws[f'C{row}'].font = normal_font
        row += 1
    
    row += 1
    
    # Account Name Analysis
    exec_ws[f'A{row}'] = 'ACCOUNT NAME ANALYSIS'
    exec_ws[f'A{row}'].font = subheader_font
    row += 1
    
    # Group by account name and calculate totals
    account_summary = data_df.groupby('Account_Name').agg({
        'Outstanding': ['sum', 'count']
    }).round(2)
    account_summary.columns = ['Total_Outstanding', 'Account_Count']
    account_summary = account_summary.sort_values('Total_Outstanding', ascending=False).head(10)
    
    exec_ws[f'A{row}'] = 'Account Name'
    exec_ws[f'B{row}'] = 'Count'
    exec_ws[f'C{row}'] = 'Total Outstanding'
    exec_ws[f'A{row}'].font = header_font
    exec_ws[f'B{row}'].font = header_font
    exec_ws[f'C{row}'].font = header_font
    exec_ws[f'A{row}'].fill = header_fill
    exec_ws[f'B{row}'].fill = header_fill
    exec_ws[f'C{row}'].fill = header_fill
    row += 1
    
    for account_name, data in account_summary.iterrows():
        exec_ws[f'A{row}'] = account_name
        exec_ws[f'B{row}'] = int(data['Account_Count'])
        exec_ws[f'C{row}'] = f'R {data["Total_Outstanding"]:,.2f}'
        exec_ws[f'A{row}'].font = normal_font
        exec_ws[f'B{row}'].font = normal_font
        exec_ws[f'C{row}'].font = normal_font
        row += 1
    
    # Adjust column widths for executive summary
    exec_ws.column_dimensions['A'].width = 35
    exec_ws.column_dimensions['B'].width = 20
    exec_ws.column_dimensions['C'].width = 20
    
    # Create detailed data sheet
    ws = wb.create_sheet(title="JUN - Quarterly Write Off P4")
    
    # Add title
    ws['A1'] = 'JUN - QUARTERLY WRITE OFF - PAGE 4 - DETAILED DATA'
    ws['A1'].font = Font(name='Calibri', size=14, bold=True, color='1F4E79')
    ws.merge_cells('A1:Q1')
    
    # Add data starting from row 3
    for r in dataframe_to_rows(data_df, index=False, header=True):
        ws.append(r)
    
    # Style the headers (row 3)
    for col in range(1, len(data_df.columns) + 1):
        cell = ws.cell(row=3, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Style the data rows
    for row in range(4, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = normal_font
            
            # Alternate row colors
            if row % 2 == 0:
                cell.fill = PatternFill(start_color='F8F9FA', end_color='F8F9FA', fill_type='solid')
            
            # Format currency columns
            if col in [4, 5, 6, 7, 8, 9, 10, 11]:  # Financial columns
                if cell.value and isinstance(cell.value, (int, float)):
                    cell.number_format = 'R #,##0.00'
    
    # Set column widths
    column_widths = {
        'A': 12,  # Account_ID
        'B': 25,  # Account_Name
        'C': 12,  # Last_Visit
        'D': 15,  # Last_Receipt_Payment
        'E': 12,  # Current
        'F': 12,  # 30_Days
        'G': 12,  # 60_Days
        'H': 12,  # 90_Days
        'I': 12,  # 120_Days
        'J': 12,  # 150_Days
        'K': 15,  # Outstanding
        'L': 30,  # Address
        'M': 8,   # Status
        'N': 12,  # Payment_Date
        'O': 15,  # ID_Number
        'P': 20,  # Claim_Number
        'Q': 10   # Claim_Status
    }
    
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width
    
    # Add auto-filter
    ws.auto_filter.ref = f"A3:Q{ws.max_row}"
    
    # Freeze panes
    ws.freeze_panes = "A4"
    
    # Add conditional formatting for Outstanding amounts
    outstanding_col = 'K'
    rule = ColorScaleRule(
        start_type='num', start_value=0, start_color='63BE7B',
        mid_type='percentile', mid_value=50, mid_color='FFEB84',
        end_type='max', end_color='F8696B'
    )
    ws.conditional_formatting.add(f'{outstanding_col}4:{outstanding_col}{ws.max_row}', rule)
    
    # Save the file
    output_file = 'page4.xlsx'
    wb.save(output_file)
    print(f"Excel file created successfully: {output_file}")
    
    # Print summary
    print(f"\nSUMMARY:")
    print(f"Total Outstanding: R {total_outstanding:,.2f}")
    print(f"Total Accounts: {total_accounts:,}")
    print(f"Average per Account: R {avg_outstanding:,.2f}")
    print(f"Total Receipt Payments: R {total_payments:,.2f}")
    print(f"Most overdue (90+ days): R {total_90_days:,.2f}")

if __name__ == "__main__":
    create_page4_excel()

import pandas as pd
import re
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime

# Helper to clean and extract columns

def extract_clean_data(csv_path):
    # Read the CSV as raw lines
    with open(csv_path, encoding='utf-8') as f:
        lines = [line.strip() for line in f if line.strip()]
    # Try to find header row (look for 'ACCOUNT NAME' and 'ID')
    header_idx = None
    for i, line in enumerate(lines):
        if 'ACCOUNT NAME' in line and 'ID' in line:
            header_idx = i
            break
    if header_idx is None:
        # fallback: look for first row with many commas
        header_idx = max(range(len(lines)), key=lambda i: lines[i].count(','))
    # Split header and data
    header = [h.strip() for h in lines[header_idx].split(',')]
    data_rows = []
    for line in lines[header_idx+1:]:
        row = [c.strip() for c in line.split(',')]
        if len(row) >= 5:  # skip empty/short rows
            data_rows.append(row)
    # Pad rows to header length
    for row in data_rows:
        while len(row) < len(header):
            row.append('')
    df = pd.DataFrame(data_rows, columns=header)
    return df

def clean_and_format(df):
    # Try to find columns for account, id, name, date, amount
    col_map = {}
    for col in df.columns:
        lcol = col.lower()
        if 'account' in lcol and 'name' not in lcol:
            col_map['Account Number'] = col
        elif 'id' in lcol:
            col_map['ID Number'] = col
        elif 'name' in lcol:
            col_map['Account Name'] = col
        elif 'date' in lcol or 'visit' in lcol:
            col_map['Date'] = col
        elif 'amount' in lcol or 'outstanding' in lcol or 'current' in lcol:
            col_map['Amount'] = col
    # Build a new DataFrame with only the important columns
    cols = ['Account Name', 'Account Number', 'ID Number', 'Date', 'Amount']
    new_df = pd.DataFrame()
    for c in cols:
        if c in col_map:
            new_df[c] = df[col_map[c]]
        else:
            new_df[c] = ''
    # Clean up dates and amounts
    if 'Date' in new_df:
        new_df['Date'] = new_df['Date'].apply(lambda x: re.sub(r'[^0-9./-]', '', x))
    if 'Amount' in new_df:
        new_df['Amount'] = new_df['Amount'].apply(lambda x: re.sub(r'[^0-9.,-]', '', x).replace(',', '.'))
    return new_df

def save_to_xlsx(df, out_path):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Accounts'
    # Title
    ws.merge_cells('A1:E1')
    ws['A1'] = 'PROMED REPORT - CLEANED & FORMATTED'
    ws['A1'].font = Font(size=14, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')
    # Subtitle
    ws.merge_cells('A2:E2')
    ws['A2'] = f'Report generated: {datetime.now().strftime("%d/%m/%Y %H:%M:%S")}'
    ws['A2'].alignment = Alignment(horizontal='center')
    # Header
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    for c_idx, col in enumerate(df.columns, 1):
        cell = ws.cell(row=3, column=c_idx, value=col)
        cell.font = header_font
        cell.fill = header_fill
    # Data
    for r_idx, row in enumerate(df.itertuples(index=False), 4):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)
    # Auto-width
    for col in ws.columns:
        max_length = 0
        col_letter = None
        for cell in col:
            if col_letter is None:
                col_letter = cell.column_letter
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        if col_letter:
            ws.column_dimensions[col_letter].width = max_length + 2
    wb.save(out_path)
    print(f'Saved cleaned XLSX: {out_path}')

if __name__ == '__main__':
    import sys
    if len(sys.argv) < 2:
        print('Usage: python clean_csv_to_xlsx.py <input.csv> [output.xlsx]')
        sys.exit(1)
    csv_path = sys.argv[1]
    out_path = sys.argv[2] if len(sys.argv) > 2 else 'cleaned_report.xlsx'
    df = extract_clean_data(csv_path)
    df2 = clean_and_format(df)
    save_to_xlsx(df2, out_path)
    print('Done. Please check the output XLSX for a professional, human-friendly layout.')

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os

def create_combined_reports_excel():
    """Create a combined Excel file with both reports1 and reports2 data"""
    try:
        # File paths
        csv_file1 = r"c:\Users\Admin\Desktop\ELC\Workspaces\Time Attendance\reports1.csv"
        csv_file2 = r"c:\Users\Admin\Desktop\ELC\Workspaces\Time Attendance\reports2.csv"
        excel_file = r"c:\Users\Admin\Desktop\ELC\Workspaces\Time Attendance\PROMED_Complete_Report.xlsx"
        
        print("Creating combined PROMED report with both pages...")
        
        def process_csv_file(csv_file):
            """Process a single CSV file and return header info and data"""
            with open(csv_file, 'r', encoding='utf-8') as f:
                lines = f.readlines()
            
            # Extract header information (first 6 lines)
            header_info = []
            for i in range(6):
                parts = lines[i].strip().split(',', 1)
                if len(parts) == 2:
                    header_info.append([parts[0], parts[1]])
                else:
                    header_info.append([parts[0], ''])
            
            # Extract data starting from line 8
            data_lines = lines[7:]
            data_content = ''.join(data_lines)
            
            from io import StringIO
            data_df = pd.read_csv(StringIO(data_content))
            
            return header_info, data_df
        
        # Process both files
        header1, data1 = process_csv_file(csv_file1)
        header2, data2 = process_csv_file(csv_file2)
        
        # Combine data
        combined_data = pd.concat([data1, data2], ignore_index=True)
        
        print(f"Page 1 records: {len(data1)}")
        print(f"Page 2 records: {len(data2)}")
        print(f"Combined records: {len(combined_data)}")
        
        # Create Excel workbook
        wb = openpyxl.Workbook()
        
        # Define styles
        title_font = Font(name='Arial', size=16, bold=True, color='000080')
        subtitle_font = Font(name='Arial', size=14, bold=True, color='000080')
        header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        data_font = Font(name='Arial', size=10)
        alt_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        
        # Create Summary sheet
        summary_sheet = wb.active
        summary_sheet.title = "Report Summary"
        
        summary_sheet['A1'] = 'PROMED MEDICAL BILLING REPORT - COMPLETE'
        summary_sheet['A1'].font = title_font
        summary_sheet.merge_cells('A1:D1')
        summary_sheet['A1'].alignment = Alignment(horizontal='center')
        
        # Add report information
        summary_sheet['A3'] = 'Report Information'
        summary_sheet['A3'].font = subtitle_font
        
        row = 4
        for key, value in header1:
            summary_sheet[f'A{row}'] = key.replace('_', ' ').title()
            summary_sheet[f'B{row}'] = value
            summary_sheet[f'A{row}'].font = Font(name='Arial', size=11, bold=True)
            row += 1
        
        # Add summary statistics
        summary_sheet[f'A{row + 1}'] = 'Summary Statistics'
        summary_sheet[f'A{row + 1}'].font = subtitle_font
        
        summary_sheet[f'A{row + 2}'] = 'Total Records (All Pages):'
        summary_sheet[f'B{row + 2}'] = len(combined_data)
        
        summary_sheet[f'A{row + 3}'] = 'Page 1 Records:'
        summary_sheet[f'B{row + 3}'] = len(data1)
        
        summary_sheet[f'A{row + 4}'] = 'Page 2 Records:'
        summary_sheet[f'B{row + 4}'] = len(data2)
        
        # Calculate total outstanding
        try:
            total_outstanding = pd.to_numeric(combined_data.iloc[:, 10], errors='coerce').sum()
            summary_sheet[f'A{row + 5}'] = 'Total Outstanding Amount:'
            summary_sheet[f'B{row + 5}'] = total_outstanding
            summary_sheet[f'B{row + 5}'].number_format = '#,##0.00'
        except:
            pass
        
        # Adjust column widths
        summary_sheet.column_dimensions['A'].width = 30
        summary_sheet.column_dimensions['B'].width = 40
        
        # Create Combined Data sheet
        data_sheet = wb.create_sheet(title="All Patient Data")
        
        data_sheet['A1'] = 'COMPLETE PATIENT BILLING DATA (ALL PAGES)'
        data_sheet['A1'].font = title_font
        data_sheet.merge_cells('A1:Q1')
        data_sheet['A1'].alignment = Alignment(horizontal='center')
        
        # Add column headers
        start_row = 3
        for col_num, column_title in enumerate(combined_data.columns, 1):
            cell = data_sheet.cell(row=start_row, column=col_num)
            cell.value = column_title.replace('_', ' ').title()
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Add data rows
        for row_num, (index, row) in enumerate(combined_data.iterrows(), start_row + 1):
            for col_num, value in enumerate(row, 1):
                cell = data_sheet.cell(row=row_num, column=col_num)
                
                # Handle numeric values for amount columns
                if col_num in [4, 5, 6, 7, 8, 9, 10]:  # Amount columns
                    try:
                        if pd.notna(value) and str(value).strip() != '':
                            cell.value = float(value)
                            cell.number_format = '#,##0.00'
                        else:
                            cell.value = 0.0
                            cell.number_format = '#,##0.00'
                    except:
                        cell.value = str(value) if pd.notna(value) else ''
                else:
                    cell.value = str(value) if pd.notna(value) else ''
                
                cell.font = data_font
                cell.alignment = Alignment(vertical='center')
                
                # Alternate row coloring
                if row_num % 2 == 0:
                    cell.fill = alt_fill
        
        # Auto-adjust column widths
        for col_num in range(1, len(combined_data.columns) + 1):
            max_length = 0
            column_letter = openpyxl.utils.get_column_letter(col_num)
            
            for row_num in range(start_row, data_sheet.max_row + 1):
                cell = data_sheet.cell(row=row_num, column=col_num)
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            data_sheet.column_dimensions[column_letter].width = adjusted_width
        
        # Add auto-filter
        last_col_letter = openpyxl.utils.get_column_letter(len(combined_data.columns))
        filter_range = f"A{start_row}:{last_col_letter}{data_sheet.max_row}"
        data_sheet.auto_filter.ref = filter_range
        
        # Freeze panes
        data_sheet.freeze_panes = f'A{start_row + 1}'
        
        # Create separate sheets for each page
        for page_num, (page_data, page_header) in enumerate([(data1, header1), (data2, header2)], 1):
            sheet = wb.create_sheet(title=f"Page {page_num} Data")
            
            sheet['A1'] = f'PATIENT BILLING DATA - PAGE {page_num}'
            sheet['A1'].font = subtitle_font
            sheet.merge_cells('A1:Q1')
            sheet['A1'].alignment = Alignment(horizontal='center')
            
            # Add headers and data (similar to above)
            start_row = 3
            for col_num, column_title in enumerate(page_data.columns, 1):
                cell = sheet.cell(row=start_row, column=col_num)
                cell.value = column_title.replace('_', ' ').title()
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            for row_num, (index, row) in enumerate(page_data.iterrows(), start_row + 1):
                for col_num, value in enumerate(row, 1):
                    cell = sheet.cell(row=row_num, column=col_num)
                    
                    if col_num in [4, 5, 6, 7, 8, 9, 10]:
                        try:
                            if pd.notna(value) and str(value).strip() != '':
                                cell.value = float(value)
                                cell.number_format = '#,##0.00'
                            else:
                                cell.value = 0.0
                                cell.number_format = '#,##0.00'
                        except:
                            cell.value = str(value) if pd.notna(value) else ''
                    else:
                        cell.value = str(value) if pd.notna(value) else ''
                    
                    cell.font = data_font
                    cell.alignment = Alignment(vertical='center')
                    
                    if row_num % 2 == 0:
                        cell.fill = alt_fill
        
        # Save the workbook
        wb.save(excel_file)
        print(f"\n✅ Combined Excel file created successfully: {excel_file}")
        
        return True
        
    except Exception as e:
        print(f"❌ Error during conversion: {str(e)}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    success = create_combined_reports_excel()
    if success:
        print("\n📊 PROMED Complete Report Created!")
        print("📄 File: PROMED_Complete_Report.xlsx")
        print("📋 Contains 4 sheets:")
        print("   1. Report Summary - Overview and statistics")
        print("   2. All Patient Data - Combined data from all pages")
        print("   3. Page 1 Data - Individual page 1 data")
        print("   4. Page 2 Data - Individual page 2 data")
        print("\n🔧 Features:")
        print("   ✓ Professional formatting and styling")
        print("   ✓ Auto-filters for easy data analysis")
        print("   ✓ Frozen header rows")
        print("   ✓ Currency formatting for amounts")
        print("   ✓ Alternating row colors")
        print("   ✓ Summary statistics")
    else:
        print("❌ Failed to create combined report!")
